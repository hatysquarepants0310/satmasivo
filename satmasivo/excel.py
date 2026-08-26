"""Excel al layout de Masiva erpDOZ: Resumen / Ingresos / Pagos."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from satmasivo.cfdi import CfdiRow, format_fecha_masiva, regimen_label

# Encabezados idénticos al Excel de Masiva (PaoloMzoJun26.xlsx).
RESUMEN_COLS = [
    "COUNT",
    "FECHA",
    "TIPO",
    "RFC_REC",
    "RECEPTOR",
    "recDomFiscal",
    "recRegimenFiscal",
    "UsoCFDI",
    "RFC_EMI",
    "EMISOR",
    "LugarExpedicion",
    "RegimenFiscal",
    "SERIE",
    "FOLIO",
    "UUID",
    "METODOPAGO",
    "NUMCTAPAGO",
    "FORMAPAGO",
    "MONEDA",
    "TIPOCAMBIO",
    "SUBTOTAL",
    "IMPUESTOSTRAS",
    "IMPUESTOSTRASIVA",
    "IMPUESTOSTRASIEPS",
    "IMPUESTOSRETE",
    "IMPUESTOSRETIVA",
    "IMPUESTOSRETISR",
    "DESCUENTOS",
    "TOTAL",
    "ImpLocalTotalRet",
    "ImpLocalTotalTras",
    "ImpLocalDetRets",
    "ImpLocalDetTras",
    "CONCEPTOS",
    "ESTADO",
    "RFCEmisorCancelado",
    "RFCEmisorCompleto",
    "RFCReceptorCancelado",
    "RFCReceptorCompleto",
]

INGRESOS_COLS = [
    "Versión",
    "UUID",
    "Serie",
    "Folio",
    "Fecha",
    "FormaPago",
    "CondicionesDePago",
    "SubTotal",
    "Descuento",
    "ImpTrasladados",
    "ImpRetenidos",
    "Total",
    "TipoCambio",
    "Moneda",
    "TipoComprobante",
    "MétodoPago",
    "LugarExpedicion",
    "Confirmacion",
    "TipoRelacion",
    "CfdiRelacionados",
    "EmisorRFC",
    "EmisorNombre",
    "RégimenFiscal",
    "ReceptorRFC",
    "ReceptorNombre",
    "ReceptorUsoCFDi",
    "ReceptorResidencia",
    "ReceptorIdTrib",
    "ClaveProdServ",
    "NoIdentificacion",
    "Cantidad",
    "ClaveUnidad",
    "Unidad",
    "Descripcion",
    "ValorUnitario",
    "Importe",
    "DescuentoConcepto",
    "BaseTrasIVATEX",
    "TrasIVATEX",
    "BaseTrasIVAT00",
    "TrasIVAT00",
    "BaseTrasIVAT08",
    "TrasIVAT08",
    "BaseTrasIVAT16",
    "TrasIVAT16",
    "BaseRetIVA",
    "RetIVA",
    "BaseRetISR",
    "RetISR",
    "Traslados",
    "Retenciones",
    "NúmPedimento",
    "CtaPredial",
]

PAGOS_COLS = [
    "UUID",
    "Folio",
    "Serie",
    "Fecha",
    "EmisorRFC",
    "Emisor",
    "EmisorRégimen",
    "ReceptorRFC",
    "Receptor",
    "FormaDePago",
    "FechaDePago",
    "NúmOperación",
    "TipoDecambio",
    "Monto",
    "Moneda",
    "NúmParcialidad",
    "UUIDpag",
    "FolioPag",
    "SeriePag",
    "MonedaPag",
    "TC Pag",
    "SaldoAnt",
    "Pago",
    "SaldoInsoluto",
]

MASIVA_TITLES = RESUMEN_COLS

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=9)
BODY_FONT = Font(name="Calibri", size=9)
MONEY = "#,##0.00"


def _blank(value):
    if value is None:
        return None
    if isinstance(value, str) and value == "":
        return None
    return value


def _money(value):
    if value is None:
        return None
    try:
        if float(value) == 0:
            return 0
    except (TypeError, ValueError):
        return value
    return float(value)


def _estado(row: CfdiRow) -> str:
    est = (row.estatus_sat or "").strip()
    if not est:
        return ""
    cod = (row.codigo_estatus or "").strip()
    if " - " in cod:
        cod = cod.split(" - ", 1)[1].strip()
    if cod.lower().startswith("comprobante") or "satisfactoriamente" in cod.lower():
        return f" {cod}/Estado:{est}"
    if est.lower().startswith("comprobante"):
        return est
    return f"Estado:{est}"


def _money_tipo(row: CfdiRow, value):
    n = _money(value)
    if n is None:
        return None
    if row.tipo_comprobante == "E" and isinstance(n, (int, float)) and n > 0:
        return -n
    return n


def _write_header(ws: Worksheet, titles: list[str]) -> None:
    ws.freeze_panes = "A2"
    for col, title in enumerate(titles, 1):
        cell = ws.cell(1, col, title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 22


def _put(ws: Worksheet, r: int, values: list, money_idx: set[int] | None = None) -> None:
    money_idx = money_idx or set()
    for c, val in enumerate(values, 1):
        cell = ws.cell(r, c, _blank(val))
        cell.font = BODY_FONT
        if c in money_idx and isinstance(val, (int, float)):
            cell.number_format = MONEY


def _resumen_row(n: int, row: CfdiRow) -> list:
    return [
        n,
        format_fecha_masiva(row.fecha),
        row.tipo_comprobante,
        row.rfc_receptor,
        row.nombre_receptor,
        row.rec_dom_fiscal,
        regimen_label(row.rec_regimen),
        row.uso_cfdi,
        row.rfc_emisor,
        row.nombre_emisor,
        row.lugar_expedicion,
        regimen_label(row.regimen_fiscal),
        row.serie,
        row.folio,
        (row.uuid or "").upper(),
        row.metodo_pago,
        row.num_cta_pago,
        row.forma_pago,
        row.moneda,
        row.tipo_cambio or None,
        _money_tipo(row, row.subtotal),
        _money_tipo(row, row.impuestos_trasladados) if row.impuestos_trasladados else None,
        _money_tipo(row, row.iva_trasladado),
        _money_tipo(row, row.ieps_trasladado),
        _money_tipo(row, row.impuestos_retenidos) if row.impuestos_retenidos else None,
        _money_tipo(row, row.iva_retenido),
        _money_tipo(row, row.isr_retenido),
        _money(row.descuento) if row.descuento else None,
        _money_tipo(row, row.total),
        _money(row.imp_local_ret),
        _money(row.imp_local_tras),
        row.imp_local_det_ret or None,
        row.imp_local_det_tras or None,
        row.texto_conceptos or None,
        _estado(row),
        "No listado",
        "No listado",
        "No listado",
        "No listado",
    ]


def _ingresos_cfdi_prefix(row: CfdiRow) -> list:
    return [
        row.version,
        (row.uuid or "").upper(),
        row.serie,
        row.folio,
        format_fecha_masiva(row.fecha),
        row.forma_pago,
        row.condiciones_pago or None,
        _money(row.subtotal),
        _money(row.descuento) if row.descuento else None,
        _money(row.impuestos_trasladados) if row.impuestos_trasladados else None,
        _money(row.impuestos_retenidos) if row.impuestos_retenidos else None,
        _money(row.total),
        row.tipo_cambio or None,
        row.moneda,
        row.tipo_comprobante,
        row.metodo_pago,
        row.lugar_expedicion,
        row.confirmacion or None,
        row.tipo_relacion or None,
        row.cfdi_relacionados or None,
        row.rfc_emisor,
        row.nombre_emisor,
        f"Item{row.regimen_fiscal}" if row.regimen_fiscal else None,
        row.rfc_receptor,
        row.nombre_receptor,
        row.uso_cfdi,
        row.rec_residencia or None,
        row.rec_num_reg_trib or None,
    ]


def _concepto_suffix(c) -> list:
    return [
        c.clave_prod,
        c.no_id or None,
        c.cantidad or None,
        c.clave_unidad or None,
        c.unidad or None,
        c.descripcion,
        _money(c.valor_unitario),
        _money(c.importe),
        _money(c.descuento),
        _money(c.base_iva_exento) or None,
        _money(c.iva_exento) or None,
        _money(c.base_iva_0) or None,
        _money(c.iva_0) or None,
        _money(c.base_iva_8) or None,
        _money(c.iva_8) or None,
        _money(c.base_iva_16) or None,
        _money(c.iva_16) or None,
        _money(c.base_ret_iva) or None,
        _money(c.ret_iva) or None,
        _money(c.base_ret_isr) or None,
        _money(c.ret_isr) or None,
        None,
        None,
        c.pedimento or None,
        c.predial or None,
    ]


def _write_resumen(ws: Worksheet, rows: list[CfdiRow]) -> None:
    _write_header(ws, RESUMEN_COLS)
    ordered = sorted(rows, key=lambda r: (r.fecha, r.uuid))
    money = {21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31}
    for i, row in enumerate(ordered, 1):
        _put(ws, i + 1, _resumen_row(i, row), money)
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["E"].width = 32
    ws.column_dimensions["J"].width = 32
    ws.column_dimensions["O"].width = 38
    ws.column_dimensions["AH"].width = 36


def _write_conceptos(ws: Worksheet, rows: list[CfdiRow], tipos: set[str]) -> None:
    _write_header(ws, INGRESOS_COLS)
    r = 2
    money = set(range(8, 13)) | set(range(35, 52))
    for row in sorted(rows, key=lambda x: (x.fecha, x.uuid)):
        if row.tipo_comprobante not in tipos:
            continue
        prefix = _ingresos_cfdi_prefix(row)
        conceptos = row.conceptos or [None]
        first = True
        for c in conceptos:
            if first:
                values = prefix + (_concepto_suffix(c) if c else [None] * 25)
                first = False
            else:
                values = [None] * 28 + _concepto_suffix(c)
            _put(ws, r, values, money)
            r += 1
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["AH"].width = 36


def _write_ingresos(ws: Worksheet, rows: list[CfdiRow]) -> None:
    _write_conceptos(ws, rows, {"I", "N", "T"})


def _write_egresos(ws: Worksheet, rows: list[CfdiRow]) -> None:
    _write_conceptos(ws, rows, {"E"})


def _write_pagos(ws: Worksheet, rows: list[CfdiRow]) -> None:
    _write_header(ws, PAGOS_COLS)
    r = 2
    money = {14, 22, 23, 24}
    for row in sorted(rows, key=lambda x: (x.fecha, x.uuid)):
        if row.tipo_comprobante != "P" and not row.pagos:
            continue
        pagos = row.pagos or [None]
        for p in pagos:
            _put(
                ws,
                r,
                [
                    (row.uuid or "").upper(),
                    row.folio,
                    row.serie,
                    format_fecha_masiva(row.fecha),
                    row.rfc_emisor,
                    row.nombre_emisor,
                    regimen_label(row.regimen_fiscal),
                    row.rfc_receptor,
                    row.nombre_receptor,
                    p.forma_pago if p else None,
                    format_fecha_masiva(p.fecha_pago) if p else None,
                    p.num_operacion if p else None,
                    p.tipo_cambio if p else None,
                    _money(p.monto) if p else None,
                    p.moneda if p else None,
                    p.num_parcialidad if p else None,
                    p.uuid_dr if p else None,
                    p.folio_dr if p else None,
                    p.serie_dr if p else None,
                    p.moneda_dr if p else None,
                    p.tc_dr if p else None,
                    _money(p.saldo_ant) if p else None,
                    _money(p.pago) if p else None,
                    _money(p.saldo_insoluto) if p else None,
                ],
                money,
            )
            r += 1
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["F"].width = 32
    ws.column_dimensions["Q"].width = 38


def resolve_xlsx_dest(path: str | Path) -> Path:
    """Cualquier ruta: absoluta, .xlsx, crea padres. Si no hay permiso, ~/satmasivo."""
    p = Path(path).expanduser()
    if not p.is_absolute():
        p = Path.home() / "satmasivo" / p.name
    if p.exists() and p.is_dir():
        p = p / "reporte.xlsx"
    if p.suffix.lower() != ".xlsx":
        p = p.with_suffix(".xlsx")
    try:
        p.parent.mkdir(parents=True, exist_ok=True)
        probe = p.parent / ".satmasivo-w"
        probe.write_text("ok", encoding="utf-8")
        probe.unlink(missing_ok=True)
        return p
    except OSError:
        fallback = Path.home() / "satmasivo" / p.name
        fallback.parent.mkdir(parents=True, exist_ok=True)
        return fallback


def export_excel(rows: list[CfdiRow], path: str | Path, rfc_firma: str | None = None) -> Path:
    del rfc_firma
    path = resolve_xlsx_dest(path)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Resumen"
    _write_resumen(ws, rows)
    tipos = {r.tipo_comprobante for r in rows}
    if tipos & {"I", "N", "T"} or not tipos:
        _write_ingresos(wb.create_sheet("Ingresos"), rows)
    if "E" in tipos:
        _write_egresos(wb.create_sheet("Egresos"), rows)
    if "P" in tipos or any(r.pagos for r in rows):
        _write_pagos(wb.create_sheet("Pagos"), rows)
    wb.save(path)
    return path
