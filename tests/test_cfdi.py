from pathlib import Path

from satmasivo.cfdi import format_fecha_masiva, parse_cfdi, scan_folder
from satmasivo.excel import RESUMEN_COLS, export_excel
from satmasivo.pdf import cfdi_to_pdf
from satmasivo.validar import expresion_impresa, parse_consulta_xml

FIXTURES = Path(__file__).parent / "fixtures"


def test_parse_ingreso():
    row = parse_cfdi(FIXTURES / "cfdi_ingreso.xml")
    assert row.uuid == "11111111-2222-3333-4444-555555555555"
    assert row.tipo_documento == "Ingreso"
    assert row.rfc_emisor == "EKU9003173C9"
    assert row.rfc_receptor == "XAXX010101000"
    assert row.iva_trasladado == 160
    assert row.total == 1160
    assert row.forma_pago == "03"
    assert row.complemento_pago == "No"
    assert row.lugar_expedicion == "55700"
    assert row.primer_concepto == "Servicio"
    assert len(row.conceptos) == 1


def test_parse_pago():
    row = parse_cfdi(FIXTURES / "cfdi_pago.xml")
    assert row.tipo_documento == "Pago"
    assert row.complemento_pago == "Sí"


def test_fecha_masiva():
    assert format_fecha_masiva("2026-03-20T14:01:40") == "20/03/2026 02:01:40 p. m."
    assert format_fecha_masiva("2026-03-20T09:06:47") == "20/03/2026 09:06:47 a. m."
    assert format_fecha_masiva("2026-07-07T00:00:00") == "07/07/2026"


def test_scan_and_excel(tmp_path):
    from openpyxl import load_workbook

    rows = scan_folder(FIXTURES)
    assert len(rows) == 2
    dest = tmp_path / "out.xlsx"
    export_excel(rows, dest, rfc_firma="EKU9003173C9")
    wb = load_workbook(dest)
    assert wb.sheetnames == ["Resumen", "Ingresos", "Pagos"]
    headers = [c.value for c in wb["Resumen"][1]]
    assert headers == RESUMEN_COLS
    assert wb["Resumen"]["C2"].value in {"I", "P"}


def test_pdf(tmp_path):
    dest = tmp_path / "f.pdf"
    cfdi_to_pdf(FIXTURES / "cfdi_ingreso.xml", dest)
    assert dest.is_file()
    assert dest.stat().st_size > 500


def test_expresion_impresa_y_soap():
    row = parse_cfdi(FIXTURES / "cfdi_ingreso.xml")
    expr = expresion_impresa(row)
    assert "re=EKU9003173C9" in expr
    assert "rr=XAXX010101000" in expr
    assert "tt=1160" in expr
    assert "id=11111111-2222-3333-4444-555555555555" in expr
    soap = """
    <s:Envelope xmlns:s="http://schemas.xmlsoap.org/soap/envelope/">
      <s:Body>
        <ConsultaResponse xmlns="http://tempuri.org/">
          <ConsultaResult xmlns:a="http://schemas.datacontract.org/2004/07/Sat.Cfdi.Negocio.ConsultaCfdi.Servicio">
            <a:CodigoEstatus>S - Comprobante obtenido satisfactoriamente.</a:CodigoEstatus>
            <a:EsCancelable>Cancelable sin aceptación</a:EsCancelable>
            <a:Estado>Vigente</a:Estado>
            <a:EstatusCancelacion/>
          </ConsultaResult>
        </ConsultaResponse>
      </s:Body>
    </s:Envelope>
    """
    data = parse_consulta_xml(soap)
    assert data["Estado"] == "Vigente"
    assert data["EsCancelable"].startswith("Cancelable")


def test_excel_estado_egreso_y_69b(tmp_path):
    from openpyxl import load_workbook

    from satmasivo.excel import _estado

    row = parse_cfdi(FIXTURES / "cfdi_ingreso.xml")
    row.estatus_sat = "Vigente"
    row.codigo_estatus = "S - Comprobante obtenido satisfactoriamente."
    est = _estado(row)
    assert "Comprobante obtenido satisfactoriamente" in est
    assert est.endswith("/Estado:Vigente")
    row.tipo_comprobante = "E"
    dest = tmp_path / "e.xlsx"
    export_excel([row], dest)
    wb = load_workbook(dest)
    assert wb.sheetnames == ["Resumen", "Egresos"]
    headers = [c.value for c in wb["Resumen"][1]]
    total_col = headers.index("TOTAL") + 1
    assert wb["Resumen"].cell(2, total_col).value == -1160
    assert wb["Resumen"]["AJ2"].value == "No listado"
    assert wb["Resumen"]["O2"].value == row.uuid.upper()


def test_export_excel_otra_ruta(tmp_path):
    dest = tmp_path / "otra" / "carpeta" / "junio"
    rows = scan_folder(FIXTURES)
    out = export_excel(rows, dest)
    assert out.is_file()
    assert out.suffix == ".xlsx"
    assert out.parent == dest.parent


